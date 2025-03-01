import openpyxl

# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("table.xlsx")

# Define variable to read the active sheet:
worksheet = wookbook.active

# Iterate the loop to read the cell values
for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, 1):
        print(col[i].value, end="\t\t")
    print('')